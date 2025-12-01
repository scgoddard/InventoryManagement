#!/usr/bin/env python3
"""
Setup Dashboard with Excel formulas to auto-calculate metrics
This makes the Dashboard dynamically update based on Gear Inventory and Checkout Log data
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


def setup_dashboard_formulas():
    """Add Excel formulas to Dashboard sheet for automatic metric calculation"""
    excel_file = '/home/user/InventoryManagement/Inventory Managment System.xlsx'

    print("Loading workbook...")
    wb = openpyxl.load_workbook(excel_file)

    # Get or create dashboard sheet
    if 'dashboard' in wb.sheetnames:
        ws = wb['dashboard']
    else:
        ws = wb.create_sheet('dashboard')

    # Clear existing data except headers
    ws.delete_rows(2, ws.max_row)

    # Set up headers
    ws['A1'] = 'Metric'
    ws['B1'] = 'Value'
    ws['C1'] = 'Last Updated'

    # Define metrics with Excel formulas
    # Note: gear_inventory data starts at row 2, checkout_log data starts at row 2
    metrics = [
        ('Total Gear Items', '=COUNTA(gear_inventory!A2:A1000)'),
        ('Available Items', '=COUNTIF(gear_inventory!E2:E1000,"Available")'),
        ('Checked Out Items', '=COUNTIF(gear_inventory!E2:E1000,"Checked Out")'),
        ('Overdue Items', '=SUMPRODUCT((gear_inventory!E2:E1000="Checked Out")*(gear_inventory!G2:G1000<TODAY())*(gear_inventory!G2:G1000<>""))'),
        ('Items in Maintenance', '=COUNTIF(gear_inventory!E2:E1000,"In Maintenance")'),
        ('Lost Items', '=COUNTIF(gear_inventory!E2:E1000,"Lost")'),
        ('Total Users', '=SUMPRODUCT(1/COUNTIF(checkout_log!D2:D1000,checkout_log!D2:D1000&""))'),
        ('Active Checkouts', '=COUNTIF(checkout_log!H2:H1000,"Active")+COUNTIF(checkout_log!H2:H1000,"Overdue")'),
        ('Total Transactions', '=COUNTA(checkout_log!A2:A1000)'),
        ('Completed Transactions', '=COUNTIF(checkout_log!H2:H1000,"Completed")'),
        ('Utilization Rate', '=IF(B2>0,B4/B2,0)'),
        ('Overdue Rate', '=IF(B4>0,B5/B4,0)'),
    ]

    # Add metrics to sheet
    for idx, (metric_name, formula) in enumerate(metrics, start=2):
        ws[f'A{idx}'] = metric_name
        ws[f'B{idx}'] = formula
        ws[f'C{idx}'] = '=TODAY()'

    # Format the sheet
    # Make headers bold
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    # Format value column as numbers with 4 decimal places
    for row in range(2, len(metrics) + 2):
        ws[f'B{row}'].number_format = '0.0000'

    # Format date column
    for row in range(2, len(metrics) + 2):
        ws[f'C{row}'].number_format = 'YYYY-MM-DD'

    print("Saving workbook with formulas...")
    wb.save(excel_file)
    print("✓ Dashboard formulas have been set up!")
    print("\nThe Dashboard will now automatically calculate metrics based on:")
    print("  - gear_inventory sheet")
    print("  - checkout_log sheet")


if __name__ == "__main__":
    setup_dashboard_formulas()
